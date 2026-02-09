
"""
file = open("sample.txt", "w")
file.write("Hello, this is a file handling example.")
file.close()

file = open("sample.txt", "r")
content = file.read()
print(content)
file.close()

try:
    with open("missing.txt", "r") as file:
        print(file.read())
except FileNotFoundError:
    print("File not found. Please check the filename.")
"""
"""
import csv
with open("students.csv", "r") as file:
    reader = csv.reader(file)
    for row in reader:
        print(row)

import csv

with open("students.csv", "r") as file:
    reader = csv.reader(file)
    
    next(reader)  
    
    for row in reader:
        print(row[0]) 
"""

from openpyxl import load_workbook

# Load existing Excel file
wb = load_workbook("students.xlsx")

# Select sheet
ws = wb.active

# Print all rows
for row in ws.iter_rows(values_only=True):
    print(row)



